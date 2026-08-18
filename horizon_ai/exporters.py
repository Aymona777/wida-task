"""
Horizon B2B Services - Data Exporters (Excel, CSV, JSON)
وحدة تصدير البيانات إلى جداول إكسل، CSV، وJSON مع دعم كامل للغة العربية
"""

import io
import csv
import json
from typing import List, Dict, Any
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class DataExporter:
    """
    Exports request records and analytics into Excel, CSV, or JSON.
    """

    HEADERS = [
        ("رمز الطلب", "request_code"),
        ("اسم الجهة", "organization_name"),
        ("شخص التواصل", "contact_person"),
        ("الصفة / المسمى", "contact_title"),
        ("وسيلة التواصل", "contact_channel"),
        ("السجل التجاري", "cr_status"),
        ("رقم السجل", "cr_number"),
        ("الخدمة الأساسية", "primary_service_name"),
        ("الخدمة الثانوية", "secondary_service_name"),
        ("الموعد المطلوب", "requested_deadline_text"),
        ("تقييم السياسات", "policy_evaluation"),
        ("حالة المراجعة البشرية", "human_review_status"),
        ("المراجع المعتمد", "reviewer_name"),
        ("البيانات الناقصة", "missing_data"),
        ("التنبيهات المهمة", "critical_alerts"),
        ("الخطوة التالية المقترحة", "suggested_next_step"),
        ("تاريخ الاستلام", "created_at")
    ]

    def export_to_csv(self, requests: List[Dict[str, Any]]) -> str:
        """
        Exports requests as UTF-8 CSV with BOM for seamless Arabic rendering in Excel.
        """
        output = io.StringIO()
        # Add UTF-8 BOM
        output.write('\ufeff')
        
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        
        # Write headers
        writer.writerow([h[0] for h in self.HEADERS])

        # Write data rows
        for req in requests:
            row = []
            for _, key in self.HEADERS:
                val = req.get(key, "")
                if isinstance(val, list):
                    val = "، ".join(val) if val else "لا توجد"
                elif val is None:
                    val = ""
                row.append(str(val))
            writer.writerow(row)

        return output.getvalue()

    def export_to_excel_bytes(self, requests: List[Dict[str, Any]]) -> bytes:
        """
        Exports requests as a beautifully styled Excel workbook (.xlsx).
        """
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV bytes
            return self.export_to_csv(requests).encode('utf-8-sig')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "سجل طلبات هورايزون"
        ws.views.sheetView[0].rightToLeft = True  # Arabic RTL Layout

        # Styles
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Corporate Royal Blue
        data_font = Font(name="Segoe UI", size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        # Status fills
        compliant_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Green
        urgent_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")    # Amber
        violation_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Red
        oos_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")       # Indigo

        # Write header row
        for col_idx, (header_label, _) in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 28

        # Write data rows
        for row_idx, req in enumerate(requests, 2):
            ws.row_dimensions[row_idx].height = 24
            for col_idx, (_, key) in enumerate(self.HEADERS, 1):
                val = req.get(key, "")
                if isinstance(val, list):
                    val = "، ".join(val) if val else "لا توجد"
                elif val is None:
                    val = ""

                cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = right_align

                # Color policy evaluation column
                if key == "policy_evaluation":
                    cell.alignment = center_align
                    if "متوافق" in str(val):
                        cell.fill = compliant_fill
                    elif "عاجل" in str(val):
                        cell.fill = urgent_fill
                    elif "مخالف" in str(val):
                        cell.fill = violation_fill
                    elif "خارج النطاق" in str(val):
                        cell.fill = oos_fill

                # Color human review status
                if key == "human_review_status":
                    cell.alignment = center_align
                    if "تمت المراجعة" in str(val):
                        cell.fill = compliant_fill
                    elif "بانتظار" in str(val):
                        cell.fill = urgent_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        return excel_stream.getvalue()

    def export_to_json(self, requests: List[Dict[str, Any]]) -> str:
        return json.dumps({
            "company": "Horizon B2B Services - شركة هورايزون لخدمات الأعمال",
            "exported_at": datetime.now().isoformat(),
            "total_count": len(requests),
            "records": requests
        }, ensure_ascii=False, indent=2)
